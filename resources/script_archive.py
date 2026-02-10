import csv
from io import BytesIO
import re
import shutil
from pathlib import Path
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from reportlab.pdfgen import canvas
from zipfile import ZipFile
from pypdf import PdfReader

resources = Path(__file__).resolve().parent
csv_path = resources / 'testik.csv'
with open(csv_path, "w", newline="", encoding="utf-8") as f:
    writer = csv.writer(f)
    writer.writerow(["id", "name", "price"])
    writer.writerow([1,"Apple",100])
    writer.writerow([2,"Orange",200])

xlsx_file = resources / 'testik2.xlsx'
wb=Workbook()
ws=wb.active
ws.title = "Eshkere"

ws.append(["ID", "Name", "Price"])
ws.append(["1", "Apple", "100"])
ws.append(["2", "Orange", "200"])
wb.save(xlsx_file)

pdf_file = resources / 'testik23.pdf'
c=canvas.Canvas(str(pdf_file))
c.drawString(100, 750, "Test PDFfile")
c.drawString(100,730,"Product list:")
c.drawString(100,710,"1. Apple - 100")
c.drawString(100,690,"2. Orange - 200")
c.save()

zip_filik = resources / 'archivcik.zip'
with ZipFile(zip_filik, 'w') as zip_file:
    zip_file.write(resources/'testik.csv', arcname='sobaka228.csv')
    zip_file.write(resources/'testik2.xlsx', arcname='sobaka2282.xlsx')
    zip_file.write(resources/'testik23.pdf', arcname='sobaka22823.pdf')

CURRENT_DIR = Path(__file__).resolve().parent
TARGET_DIR = CURRENT_DIR.parent / 'archivedirectory'
TARGET_DIR.mkdir(exist_ok=True)
shutil.move(zip_filik, TARGET_DIR/zip_filik.name)

zip_path = TARGET_DIR/'archivcik.zip'
with ZipFile(zip_path, 'r') as zip_file:
    assert set(zip_file.namelist())=={'sobaka228.csv', 'sobaka2282.xlsx', 'sobaka22823.pdf'}

    with zip_file.open('sobaka228.csv') as f:
        reader = csv.reader(f.read().decode('utf-8').splitlines())
        rows = list(reader)
        assert rows[0]==["id", "name", "price"]
        assert len(rows)==3
        assert rows[1]==["1","Apple","100"]
        assert rows[2]==["2","Orange","200"]

    with zip_file.open('sobaka22823.pdf') as f:
        text = ''.join(page.extract_text() for page in PdfReader(f).pages)
        narmal_text = re.sub(r"\s+", " ", text)
        assert "2. Orange - 200" in text

    with zip_file.open('sobaka2282.xlsx') as f:
        wb = load_workbook(BytesIO(f.read()))
        sheet = wb.active
        assert sheet['A1'].value=='ID'
        assert sheet['A2'].value=='1'
        assert sheet['A3'].value=='2'
        assert sheet['B1'].value=='Name'
        assert sheet['C1'].value=='Price'
        assert sheet['B2'].value=='Apple'
        assert sheet['C2'].value=='100'
        assert sheet['B3'].value=='Orange'
        assert sheet['C3'].value=='200'
        print('Всё крутяк')